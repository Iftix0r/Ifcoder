from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Q, Sum
from django.db.models.functions import TruncMonth
from django.http import JsonResponse
from django.shortcuts import render
from django.urls import reverse, reverse_lazy
from django.utils import timezone
from django.views.decorators.http import require_POST
from django.views.generic import CreateView, DeleteView, DetailView, ListView, UpdateView

import html
from dashboard.mixins import CSVExportMixin
from bots.telegram import send_telegram_message

from .forms import ExpenseForm, IncomeForm, InvoiceForm
from .models import Expense, Income, Invoice


def _months_back(base, n):
    year, month = base.year, base.month - n
    while month <= 0:
        month += 12
        year -= 1
    return base.replace(year=year, month=month, day=1)


@login_required
def home(request):
    today = timezone.localdate()
    month_start = today.replace(day=1)

    month_income = Income.objects.filter(date__gte=month_start).aggregate(total=Sum("amount"))[
        "total"
    ] or 0
    month_expense = Expense.objects.filter(date__gte=month_start).aggregate(
        total=Sum("amount")
    )["total"] or 0

    outstanding_invoices = Invoice.objects.exclude(status=Invoice.Status.PAID).select_related(
        "client"
    )
    outstanding_total = outstanding_invoices.aggregate(total=Sum("amount"))["total"] or 0

    range_start = _months_back(month_start, 5)
    income_by_month = {
        row["month"]: row["total"]
        for row in Income.objects.filter(date__gte=range_start)
        .annotate(month=TruncMonth("date"))
        .values("month")
        .annotate(total=Sum("amount"))
    }
    expense_by_month = {
        row["month"]: row["total"]
        for row in Expense.objects.filter(date__gte=range_start)
        .annotate(month=TruncMonth("date"))
        .values("month")
        .annotate(total=Sum("amount"))
    }
    monthly_summary = []
    for i in range(5, -1, -1):
        m = _months_back(month_start, i)
        m_income = income_by_month.get(m, 0)
        m_expense = expense_by_month.get(m, 0)
        monthly_summary.append(
            {
                "month": m,
                "income": m_income,
                "expense": m_expense,
                "diff": m_income - m_expense,
            }
        )

    context = {
        "month_income": month_income,
        "month_expense": month_expense,
        "outstanding_invoices": outstanding_invoices[:10],
        "outstanding_count": outstanding_invoices.count(),
        "outstanding_total": outstanding_total,
        "monthly_summary": monthly_summary,
        "recent_income": Income.objects.select_related("client", "project")[:5],
        "recent_expense": Expense.objects.all()[:5],
    }
    return render(request, "finance/home.html", context)


class IncomeListView(LoginRequiredMixin, CSVExportMixin, ListView):
    model = Income
    template_name = "finance/income_list.html"
    context_object_name = "incomes"
    paginate_by = 20
    csv_filename = "daromadlar.csv"
    csv_headers = ["Summa", "Usul", "Mijoz", "Loyiha", "Sana"]

    def get_csv_row(self, obj):
        return [obj.amount, obj.get_method_display(), obj.client, obj.project, obj.date]

    def get_queryset(self):
        qs = super().get_queryset().select_related("client", "project")
        q = self.request.GET.get("q")
        if q:
            qs = qs.filter(Q(description__icontains=q) | Q(client__name__icontains=q))
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["q"] = self.request.GET.get("q", "")
        ctx["total"] = self.get_queryset().aggregate(total=Sum("amount"))["total"] or 0
        return ctx


class IncomeDetailView(LoginRequiredMixin, DetailView):
    model = Income
    template_name = "finance/income_detail.html"
    context_object_name = "income"


class IncomeCreateView(LoginRequiredMixin, CreateView):
    model = Income
    form_class = IncomeForm
    template_name = "finance/income_form.html"
    success_url = reverse_lazy("finance:income_list")

    def form_valid(self, form):
        response = super().form_valid(form)
        try:
            client_name = html.escape(str(self.object.client or "Noma'lum"))
            msg = (
                f"💵 <b>YANGI DAROMAD KELIB TUSHDI!</b>\n\n"
                f"💰 <b>Summa:</b> {self.object.amount:,.2f} UZS\n"
                f"👤 <b>Mijoz:</b> {client_name}\n"
                f"💳 <b>To'lov usuli:</b> {self.object.get_method_display()}\n"
                f"📅 <b>Sana:</b> {self.object.date}"
            )
            send_telegram_message(msg)
        except Exception:
            pass
        return response


class IncomeUpdateView(LoginRequiredMixin, UpdateView):
    model = Income
    form_class = IncomeForm
    template_name = "finance/income_form.html"

    def get_success_url(self):
        return reverse("finance:income_detail", args=[self.object.pk])


class IncomeDeleteView(LoginRequiredMixin, DeleteView):
    model = Income
    template_name = "dashboard/confirm_delete.html"
    success_url = reverse_lazy("finance:income_list")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["cancel_url"] = self.success_url
        return ctx


class ExpenseListView(LoginRequiredMixin, CSVExportMixin, ListView):
    model = Expense
    template_name = "finance/expense_list.html"
    context_object_name = "expenses"
    paginate_by = 20
    csv_filename = "xarajatlar.csv"
    csv_headers = ["Summa", "Toifa", "Sana"]

    def get_csv_row(self, obj):
        return [obj.amount, obj.get_category_display(), obj.date]

    def get_queryset(self):
        qs = super().get_queryset().select_related("project")
        q = self.request.GET.get("q")
        if q:
            qs = qs.filter(Q(description__icontains=q))
        category = self.request.GET.get("category")
        if category:
            qs = qs.filter(category=category)
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["q"] = self.request.GET.get("q", "")
        ctx["category"] = self.request.GET.get("category", "")
        ctx["category_choices"] = Expense.Category.choices
        ctx["total"] = self.get_queryset().aggregate(total=Sum("amount"))["total"] or 0
        return ctx


class ExpenseDetailView(LoginRequiredMixin, DetailView):
    model = Expense
    template_name = "finance/expense_detail.html"
    context_object_name = "expense"

    def get_queryset(self):
        return super().get_queryset().select_related("project")


class ExpenseCreateView(LoginRequiredMixin, CreateView):
    model = Expense
    form_class = ExpenseForm
    template_name = "finance/expense_form.html"
    success_url = reverse_lazy("finance:expense_list")


class ExpenseUpdateView(LoginRequiredMixin, UpdateView):
    model = Expense
    form_class = ExpenseForm
    template_name = "finance/expense_form.html"

    def get_success_url(self):
        return reverse("finance:expense_detail", args=[self.object.pk])


class ExpenseDeleteView(LoginRequiredMixin, DeleteView):
    model = Expense
    template_name = "dashboard/confirm_delete.html"
    success_url = reverse_lazy("finance:expense_list")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["cancel_url"] = self.success_url
        return ctx


class InvoiceListView(LoginRequiredMixin, CSVExportMixin, ListView):
    model = Invoice
    template_name = "finance/invoice_list.html"
    context_object_name = "invoices"
    paginate_by = 20
    csv_filename = "hisob-fakturalar.csv"
    csv_headers = ["Raqami", "Mijoz", "Summa", "Chiqarilgan sana", "Muddat", "Holati"]

    def get_csv_row(self, obj):
        return [
            obj.number,
            obj.client,
            obj.amount,
            obj.issued_date,
            obj.due_date,
            obj.get_status_display(),
        ]

    def get_queryset(self):
        qs = super().get_queryset().select_related("client", "project")
        q = self.request.GET.get("q")
        if q:
            qs = qs.filter(Q(number__icontains=q) | Q(client__name__icontains=q))
        status = self.request.GET.get("status")
        if status:
            qs = qs.filter(status=status)
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["q"] = self.request.GET.get("q", "")
        ctx["status"] = self.request.GET.get("status", "")
        ctx["status_choices"] = Invoice.Status.choices
        return ctx


class InvoiceDetailView(LoginRequiredMixin, DetailView):
    model = Invoice
    template_name = "finance/invoice_detail.html"
    context_object_name = "invoice"


class InvoiceCreateView(LoginRequiredMixin, CreateView):
    model = Invoice
    form_class = InvoiceForm
    template_name = "finance/invoice_form.html"
    success_url = reverse_lazy("finance:invoice_list")

    def form_valid(self, form):
        response = super().form_valid(form)
        try:
            client_name = html.escape(str(self.object.client))
            msg = (
                f"📄 <b>Yangi Hisob-Faktura (Invoys) Yaratildi</b>\n\n"
                f"<b>Invoys:</b> №{self.object.number}\n"
                f"👤 <b>Mijoz:</b> {client_name}\n"
                f"💰 <b>Summa:</b> {self.object.amount:,.2f} UZS\n"
                f"📅 <b>To'lov muddati:</b> {self.object.due_date}"
            )
            send_telegram_message(msg)
        except Exception:
            pass
        return response


class InvoiceUpdateView(LoginRequiredMixin, UpdateView):
    model = Invoice
    form_class = InvoiceForm
    template_name = "finance/invoice_form.html"

    def get_success_url(self):
        return reverse("finance:invoice_detail", args=[self.object.pk])


class InvoiceDeleteView(LoginRequiredMixin, DeleteView):
    model = Invoice
    template_name = "dashboard/confirm_delete.html"
    success_url = reverse_lazy("finance:invoice_list")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["cancel_url"] = self.success_url
        return ctx


class InvoicePrintView(LoginRequiredMixin, DetailView):
    model = Invoice
    template_name = "finance/invoice_print.html"
    context_object_name = "invoice"


@login_required
@require_POST
def invoice_mark_paid(request, pk):
    invoice = Invoice.objects.get(pk=pk)
    invoice.status = Invoice.Status.PAID
    invoice.save(update_fields=["status"])
    try:
        msg = (
            f"✅ <b>INVOYS TO'LANDI!</b>\n\n"
            f"📄 <b>Invoys:</b> №{invoice.number}\n"
            f"👤 <b>Mijoz:</b> {html.escape(str(invoice.client))}\n"
            f"💰 <b>Summa:</b> {invoice.amount:,.2f} UZS"
        )
        send_telegram_message(msg)
    except Exception:
        pass
    return JsonResponse({"ok": True, "label": invoice.get_status_display()})


@login_required
def invoice_pdf(request, pk):
    """Invoys uchun PDF fayl generatsiya qilib yuboradi."""
    import io
    from django.http import FileResponse
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_RIGHT, TA_CENTER

    invoice = get_object_or_404(Invoice, pk=pk)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    accent = colors.HexColor("#6366f1")
    dark = colors.HexColor("#0f1629")
    muted = colors.HexColor("#94a3b8")

    title_style = ParagraphStyle("Title", parent=styles["Title"], textColor=accent,
                                 fontSize=26, spaceAfter=4)
    sub_style = ParagraphStyle("Sub", parent=styles["Normal"], textColor=muted, fontSize=10)
    normal = styles["Normal"]
    right_style = ParagraphStyle("Right", parent=styles["Normal"], alignment=TA_RIGHT)

    elements = []
    # Header
    elements.append(Paragraph("IFCODER", title_style))
    elements.append(Paragraph("iftix0r.uz | +998 XX XXX XX XX", sub_style))
    elements.append(Spacer(1, 0.5*cm))

    # Invoice number + dates
    info_data = [
        [Paragraph("<b>Hisob-Faktura (Invoys)</b>", normal),
         Paragraph(f"<b>№ {invoice.number}</b>", right_style)],
        [Paragraph(f"Chiqarilgan: {invoice.issued_date.strftime('%d.%m.%Y')}", sub_style),
         Paragraph(f"Muddat: {invoice.due_date.strftime('%d.%m.%Y')}", right_style)],
    ]
    info_table = Table(info_data, colWidths=[9*cm, 8*cm])
    info_table.setStyle(TableStyle([
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 0.4*cm))

    # Divider line
    divider = Table([[""]], colWidths=[17*cm])
    divider.setStyle(TableStyle([
        ("LINEBELOW", (0, 0), (-1, -1), 1.5, accent),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
    ]))
    elements.append(divider)
    elements.append(Spacer(1, 0.5*cm))

    # Client info
    elements.append(Paragraph("<b>Mijoz:</b>", normal))
    elements.append(Paragraph(str(invoice.client), styles["Normal"]))
    if invoice.client.phone:
        elements.append(Paragraph(f"Tel: {invoice.client.phone}", sub_style))
    if invoice.client.email:
        elements.append(Paragraph(f"Email: {invoice.client.email}", sub_style))
    elements.append(Spacer(1, 0.5*cm))

    # Items table
    items_data = [
        [Paragraph("<b>Tavsif</b>", normal),
         Paragraph("<b>Summa</b>", right_style)],
        [invoice.notes or (f"Loyiha: {invoice.project}" if invoice.project else "Xizmatlar uchun to'lov"),
         Paragraph(f"<b>{invoice.amount:,.0f} UZS</b>", right_style)],
    ]
    items_table = Table(items_data, colWidths=[13*cm, 4*cm])
    items_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), dark),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, 0), 11),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ("RIGHTPADDING", (0, 0), (-1, -1), 10),
    ]))
    elements.append(items_table)
    elements.append(Spacer(1, 0.4*cm))

    # Total
    total_data = [["", Paragraph(f"<b>JAMI: {invoice.amount:,.0f} UZS</b>",
                                  ParagraphStyle("Total", parent=right_style, fontSize=14, textColor=accent))]]
    total_table = Table(total_data, colWidths=[13*cm, 4*cm])
    total_table.setStyle(TableStyle([("ALIGN", (1, 0), (1, -1), "RIGHT")]))
    elements.append(total_table)
    elements.append(Spacer(1, 0.8*cm))

    # Status badge area
    status_color = colors.HexColor("#22c55e") if invoice.status == "paid" else colors.HexColor("#f59e0b")
    status_data = [[Paragraph(f"Holat: <b>{invoice.get_status_display()}</b>", normal)]]
    status_table = Table(status_data, colWidths=[17*cm])
    status_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), status_color),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.white),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING", (0, 0), (-1, -1), 12),
        ("ROUNDEDCORNERS", [4]),
    ]))
    elements.append(status_table)
    elements.append(Spacer(1, 2*cm))
    elements.append(Paragraph("Rahmat! Hamkorlik uchun minnatdormiz.", sub_style))

    doc.build(elements)
    buf.seek(0)
    filename = f"invoys-{invoice.number}.pdf"
    return FileResponse(buf, as_attachment=True, filename=filename, content_type="application/pdf")


@login_required
def invoices_excel(request):
    """Barcha invoyslari Excel (xlsx) fayli sifatida yuklab olish."""
    import io
    from django.http import HttpResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    invoices = Invoice.objects.select_related("client", "project").all()
    # Apply same filters as list view
    q = request.GET.get("q")
    status = request.GET.get("status")
    from django.db.models import Q as DQ
    if q:
        invoices = invoices.filter(DQ(number__icontains=q) | DQ(client__name__icontains=q))
    if status:
        invoices = invoices.filter(status=status)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoyslari"

    # Styling
    header_fill = PatternFill("solid", fgColor="6366F1")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin_border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )

    headers = ["Raqami", "Mijoz", "Loyiha", "Summa (UZS)", "Chiqarilgan", "Muddat", "Holati"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    ws.row_dimensions[1].height = 22

    for row_idx, inv in enumerate(invoices, 2):
        row_data = [
            inv.number,
            str(inv.client),
            str(inv.project) if inv.project else "—",
            float(inv.amount),
            inv.issued_date.strftime("%d.%m.%Y"),
            inv.due_date.strftime("%d.%m.%Y"),
            inv.get_status_display(),
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx == 4:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")
            if row_idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F8FAFC")

    # Column widths
    for col, width in zip(range(1, 8), [16, 28, 22, 18, 14, 14, 16]):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="invoyslari.xlsx"'
    return response
